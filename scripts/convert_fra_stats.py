"""
Convert data/Statistics_FRA_Claims.xlsx into frontend JSON for the FRA Stats tab.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "Statistics_FRA_Claims.xlsx"
OUT = ROOT / "frontend" / "src" / "data" / "fraClaimsStatistics.json"

SHEET_META = {
    "June_2024": {"label": "June 2024", "asOf": "2024-06-30"},
    "OCT_2023": {"label": "October 2023", "asOf": "2023-10-31"},
    "NOV_2022": {"label": "November 2022", "asOf": "2022-11-30"},
    "May_2019": {"label": "May 2019", "asOf": "2019-05-31"},
}

# Map fuzzy header fragments -> stable keys (checked in order of specificity)
HEADER_RULES: list[tuple[tuple[str, ...], str]] = [
    (("claims received", "individual"), "claimsReceivedIndividual"),
    (("claims received", "community"), "claimsReceivedCommunity"),
    (("claims received", "total"), "claimsReceivedTotal"),
    (("number of claims received", "individual"), "claimsReceivedIndividual"),
    (("number of claims received", "community"), "claimsReceivedCommunity"),
    (("number of claims received", "total"), "claimsReceivedTotal"),
    (("no. of claims received", "individual"), "claimsReceivedIndividual"),
    (("no. of claims received", "community"), "claimsReceivedCommunity"),
    (("titles distributed", "individual"), "titlesDistributedIndividual"),
    (("titles distributed", "community"), "titlesDistributedCommunity"),
    (("titles distributed", "total"), "titlesDistributedTotal"),
    (("no. of titles distributed", "individual"), "titlesDistributedIndividual"),
    (("no. of titles distributed", "community"), "titlesDistributedCommunity"),
    (("pending claims", "individual"), "pendingIndividual"),
    (("pending claims", "community"), "pendingCommunity"),
    (("claims rejected", "individual"), "rejectedIndividual"),
    (("claims rejected", "community"), "rejectedCommunity"),
    (("claims rejected", "total"), "rejectedTotal"),
    (("forest land", "individual"), "forestLandIndividual"),
    (("forest land", "community"), "forestLandCommunity"),
    (("forest land", "total"), "forestLandTotal"),
    (("extent of forest land", "individual"), "forestLandIndividual"),
    (("extent of forest land", "community"), "forestLandCommunity"),
]


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def map_header(header: str) -> str | None:
    if header in {"sl. no.", "s.no.", "s. no.", "sl no.", "sno"}:
        return "slNo"
    if header in {"state", "states", "states/ut", "state/ut"}:
        return "state"

    for needles, key in HEADER_RULES:
        if all(n in header for n in needles):
            return key

    # Fallbacks when sheet has only Individual/Community without Total wording
    if "claims received" in header and "individual" in header:
        return "claimsReceivedIndividual"
    if "claims received" in header and "community" in header:
        return "claimsReceivedCommunity"
    return None


def parse_number(value: object) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    text = str(value).strip()
    if not text or text.upper() in {"NA", "NA/NR", "NR", "N/A", "-"}:
        return None

    # Dirty values like "-117*"
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if cleaned in {"", "-", ".", "-."}:
        return None
    try:
        num = float(cleaned)
        return int(num) if num.is_integer() else num
    except ValueError:
        return None


def convert_sheet(ws, sheet_id: str) -> dict:
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError(f"Sheet {sheet_id} is empty")

    keys: list[str | None] = [map_header(normalize_header(h)) for h in header_row]
    data_rows: list[dict] = []

    for raw in rows_iter:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue

        mapped: dict[str, object] = {}
        for idx, key in enumerate(keys):
            if not key or idx >= len(raw):
                continue
            val = raw[idx]
            if key == "state":
                state = str(val).strip() if val is not None else ""
                mapped["state"] = state
            elif key == "slNo":
                mapped["slNo"] = parse_number(val)
            else:
                mapped[key] = parse_number(val)

        state = str(mapped.get("state") or "").strip()
        if not state or state.lower() == "total":
            continue

        # Prefer computed totals when missing
        if mapped.get("claimsReceivedTotal") is None:
            ind = mapped.get("claimsReceivedIndividual")
            com = mapped.get("claimsReceivedCommunity")
            if isinstance(ind, (int, float)) or isinstance(com, (int, float)):
                mapped["claimsReceivedTotal"] = (ind or 0) + (com or 0)

        if mapped.get("titlesDistributedTotal") is None:
            ind = mapped.get("titlesDistributedIndividual")
            com = mapped.get("titlesDistributedCommunity")
            if isinstance(ind, (int, float)) or isinstance(com, (int, float)):
                mapped["titlesDistributedTotal"] = (ind or 0) + (com or 0)

        data_rows.append(mapped)

    meta = SHEET_META.get(sheet_id, {"label": sheet_id, "asOf": None})
    return {
        "id": sheet_id,
        "label": meta["label"],
        "asOf": meta["asOf"],
        "rows": data_rows,
    }


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")

    wb = load_workbook(XLSX, data_only=True, read_only=True)
    # Prefer newest period first
    preferred = ["June_2024", "OCT_2023", "NOV_2022", "May_2019"]
    sheet_order = [s for s in preferred if s in wb.sheetnames]
    sheet_order += [s for s in wb.sheetnames if s not in sheet_order]

    periods = [convert_sheet(wb[name], name) for name in sheet_order]
    payload = {
        "source": "Statistics_FRA_Claims.xlsx",
        "periods": periods,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {OUT} ({sum(len(p['rows']) for p in periods)} state rows across {len(periods)} periods)")


if __name__ == "__main__":
    main()
